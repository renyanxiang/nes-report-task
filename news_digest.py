#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import os
import re
import socket
import smtplib
import sys
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import feedparser
from openpyxl import Workbook
from openpyxl.styles import Font


USER_AGENT = "Mozilla/5.0 (NewsDigestBot/1.0; +https://example.com)"
DEFAULT_TIMEZONE = "Asia/Shanghai"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "output"
MAX_ITEMS_PER_FEED = 12
REQUEST_TIMEOUT_SECONDS = 15

STOPWORDS = {
    "about", "after", "again", "amid", "also", "been", "before", "being",
    "between", "could", "first", "from", "have", "into", "just", "more",
    "most", "must", "over", "says", "said", "than", "that", "their", "them",
    "they", "this", "those", "through", "today", "under", "very", "what",
    "when", "where", "which", "while", "with", "would", "world", "years",
    "year", "news", "live", "latest", "update", "updates", "analysis",
}


@dataclass
class FeedSource:
    name: str
    url: str


@dataclass
class DigestResult:
    output_path: Path
    generated_at: datetime
    items: list[dict]
    topic_summary: list[tuple[str, int]]


@dataclass
class EmailSettings:
    smtp_host: str
    smtp_port: int
    smtp_user: str
    smtp_password: str
    sender: str
    recipients: list[str]
    use_ssl: bool


@dataclass
class TranslationSettings:
    provider: str
    api_key: str
    credentials_path: str
    project_id: str
    model: str
    batch_size: int


FEEDS = [
    FeedSource("BBC World", "http://feeds.bbci.co.uk/news/world/rss.xml"),
    FeedSource("CNN World", "http://rss.cnn.com/rss/edition_world.rss"),
    FeedSource("The Guardian World", "https://www.theguardian.com/world/rss"),
    FeedSource("Al Jazeera", "https://www.aljazeera.com/xml/rss/all.xml"),
    FeedSource("DW Top Stories", "https://rss.dw.com/rdf/rss-en-top"),
    FeedSource("Financial Times World", "https://www.ft.com/world?format=rss"),
    FeedSource("New York Times World", "https://rss.nytimes.com/services/xml/rss/nyt/World.xml"),
    FeedSource("France 24", "https://www.france24.com/en/rss"),
]

socket.setdefaulttimeout(REQUEST_TIMEOUT_SECONDS)


def strip_html(text: str) -> str:
    cleaned = re.sub(r"<[^>]+>", " ", text or "")
    cleaned = html.unescape(cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def parse_datetime(entry: dict, fallback_tz: ZoneInfo) -> datetime | None:
    parsed = entry.get("published_parsed") or entry.get("updated_parsed")
    if not parsed:
        return None
    dt = datetime(*parsed[:6], tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(fallback_tz)


def fetch_feed(source: FeedSource, tz: ZoneInfo) -> list[dict]:
    feed = feedparser.parse(source.url, agent=USER_AGENT)
    if getattr(feed, "bozo", False) and not getattr(feed, "entries", None):
        print(f"[WARN] failed to parse feed: {source.name}", file=sys.stderr)
        return []

    items = []
    for entry in feed.entries[:MAX_ITEMS_PER_FEED]:
        title = strip_html(entry.get("title", "")).strip()
        summary = strip_html(entry.get("summary", "")).strip()
        link = entry.get("link", "").strip()
        published = parse_datetime(entry, tz)
        items.append(
            {
                "source": source.name,
                "title": title,
                "title_zh": "",
                "summary": summary,
                "summary_zh": "",
                "link": link,
                "published_at": published.strftime("%Y-%m-%d %H:%M:%S %Z") if published else "",
                "published_sort": published or datetime.min.replace(tzinfo=tz),
            }
        )
    return items


def tokenize(text: str) -> Iterable[str]:
    for word in re.findall(r"[A-Za-z][A-Za-z-']{2,}", text.lower()):
        if word not in STOPWORDS and not word.isdigit():
            yield word


def build_topic_summary(items: list[dict], top_n: int = 15) -> list[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for item in items:
        counter.update(tokenize(item["title"]))
    return counter.most_common(top_n)


def autosize_sheet(sheet) -> None:
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 50)


def write_excel(items: list[dict], topic_summary: list[tuple[str, int]], output_path: Path, generated_at: datetime) -> None:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Generated At", generated_at.strftime("%Y-%m-%d %H:%M:%S %Z")])
    summary_sheet.append(["Article Count", len(items)])
    summary_sheet.append([])
    summary_sheet.append(["Top Topics", "Mentions"])
    for cell in summary_sheet[4]:
        cell.font = Font(bold=True)
    for topic, count in topic_summary:
        summary_sheet.append([topic, count])

    source_sheet = workbook.create_sheet("By Source")
    source_sheet.append(["Source", "Article Count"])
    for cell in source_sheet[1]:
        cell.font = Font(bold=True)
    source_counter = Counter(item["source"] for item in items)
    for source, count in source_counter.most_common():
        source_sheet.append([source, count])

    article_sheet = workbook.create_sheet("Articles")
    headers = ["Source", "Published At", "Title (EN)", "标题（中文）", "Summary (EN)", "摘要（中文）", "Link"]
    article_sheet.append(headers)
    for cell in article_sheet[1]:
        cell.font = Font(bold=True)

    sorted_items = sorted(items, key=lambda item: item["published_sort"], reverse=True)
    for item in sorted_items:
        article_sheet.append(
            [
                item["source"],
                item["published_at"],
                item["title"],
                item.get("title_zh", ""),
                item["summary"],
                item.get("summary_zh", ""),
                item["link"],
            ]
        )

    for sheet in workbook.worksheets:
        autosize_sheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def run_once(output_dir: Path, timezone_name: str) -> DigestResult:
    tz = ZoneInfo(timezone_name)
    generated_at = datetime.now(tz)

    all_items: list[dict] = []
    for source in FEEDS:
        all_items.extend(fetch_feed(source, tz))

    topic_summary = build_topic_summary(all_items)
    output_path = output_dir / f"international_news_digest_{generated_at.strftime('%Y%m%d')}.xlsx"
    write_excel(all_items, topic_summary, output_path, generated_at)
    return DigestResult(
        output_path=output_path,
        generated_at=generated_at,
        items=all_items,
        topic_summary=topic_summary,
    )


def parse_bool(value: str | None, default: bool) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def load_translation_settings() -> TranslationSettings | None:
    provider = os.getenv("NEWS_DIGEST_TRANSLATION_PROVIDER", "google").strip().lower()
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    credentials_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    project_id = os.getenv("GOOGLE_CLOUD_PROJECT", "").strip()
    model = os.getenv("NEWS_DIGEST_OPENAI_MODEL", "gpt-4o-mini").strip()
    batch_size = os.getenv("NEWS_DIGEST_TRANSLATION_BATCH_SIZE", "8").strip()

    if provider == "google" and not credentials_path:
        return None

    if provider == "openai" and not api_key:
        return None

    return TranslationSettings(
        provider=provider,
        api_key=api_key,
        credentials_path=credentials_path,
        project_id=project_id,
        model=model,
        batch_size=max(1, int(batch_size)),
    )


def chunked(items: list[dict], chunk_size: int) -> Iterable[list[dict]]:
    for index in range(0, len(items), chunk_size):
        yield items[index:index + chunk_size]


def translate_items_to_chinese_openai(items: list[dict], settings: TranslationSettings) -> None:
    from openai import OpenAI

    client = OpenAI(api_key=settings.api_key)
    schema = {
        "type": "json_schema",
        "name": "news_digest_translation",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "translations": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "index": {"type": "integer"},
                            "title_zh": {"type": "string"},
                            "summary_zh": {"type": "string"},
                        },
                        "required": ["index", "title_zh", "summary_zh"],
                        "additionalProperties": False,
                    },
                }
            },
            "required": ["translations"],
            "additionalProperties": False,
        },
    }

    for batch in chunked(items, settings.batch_size):
        payload = [
            {
                "index": idx,
                "title": item["title"],
                "summary": item["summary"],
            }
            for idx, item in enumerate(batch)
        ]
        response = client.responses.create(
            model=settings.model,
            input=[
                {
                    "role": "system",
                    "content": (
                        "You are a professional news translator. Translate English news titles and summaries "
                        "into accurate Simplified Chinese. Preserve named entities, dates, numbers, and tone. "
                        "Return JSON only."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        "Translate the following news items into Simplified Chinese and keep them concise.\n"
                        f"{json.dumps(payload, ensure_ascii=False)}"
                    ),
                },
            ],
            text={"format": schema},
        )
        parsed = json.loads(response.output_text)
        translations = {item["index"]: item for item in parsed["translations"]}
        for idx, item in enumerate(batch):
            translated = translations.get(idx)
            if translated is None:
                continue
            item["title_zh"] = translated["title_zh"].strip()
            item["summary_zh"] = translated["summary_zh"].strip()


def translate_items_to_chinese_google(items: list[dict], settings: TranslationSettings) -> None:
    from google.cloud import translate_v2 as translate

    if settings.credentials_path:
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = settings.credentials_path
    if settings.project_id:
        os.environ["GOOGLE_CLOUD_PROJECT"] = settings.project_id

    client = translate.Client()
    texts: list[str] = []
    for item in items:
        texts.append(item["title"])
        texts.append(item["summary"])

    for start in range(0, len(texts), settings.batch_size * 2):
        batch = texts[start:start + settings.batch_size * 2]
        translations = client.translate(batch, target_language="zh-CN", source_language="en", format_="text")
        if isinstance(translations, dict):
            translations = [translations]
        for offset, translated in enumerate(translations):
            item_index = (start + offset) // 2
            is_title = ((start + offset) % 2) == 0
            translated_text = strip_html(translated.get("translatedText", ""))
            if is_title:
                items[item_index]["title_zh"] = translated_text
            else:
                items[item_index]["summary_zh"] = translated_text


def translate_items_to_chinese(items: list[dict], settings: TranslationSettings) -> None:
    if settings.provider == "google":
        translate_items_to_chinese_google(items, settings)
        return
    if settings.provider == "openai":
        translate_items_to_chinese_openai(items, settings)
        return
    raise RuntimeError(f"unsupported translation provider: {settings.provider}")


def translation_provider_label(settings: TranslationSettings) -> str:
    if settings.provider == "google":
        return "google-cloud-translate"
    if settings.provider == "openai":
        return f"openai:{settings.model}"
    return settings.provider


def load_email_settings(default_recipient: str) -> EmailSettings | None:
    smtp_host = os.getenv("NEWS_DIGEST_SMTP_HOST", "").strip()
    smtp_port = os.getenv("NEWS_DIGEST_SMTP_PORT", "465").strip()
    smtp_user = os.getenv("NEWS_DIGEST_SMTP_USER", "").strip()
    smtp_password = os.getenv("NEWS_DIGEST_SMTP_PASSWORD", "").strip()
    sender = os.getenv("NEWS_DIGEST_SENDER", smtp_user).strip()
    recipients_raw = os.getenv("NEWS_DIGEST_RECIPIENTS", default_recipient).strip()
    use_ssl = parse_bool(os.getenv("NEWS_DIGEST_SMTP_SSL"), True)

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, sender, recipients_raw]):
        return None

    recipients = [item.strip() for item in recipients_raw.split(",") if item.strip()]
    return EmailSettings(
        smtp_host=smtp_host,
        smtp_port=int(smtp_port),
        smtp_user=smtp_user,
        smtp_password=smtp_password,
        sender=sender,
        recipients=recipients,
        use_ssl=use_ssl,
    )


def build_email_html(result: DigestResult) -> str:
    sorted_items = sorted(result.items, key=lambda item: item["published_sort"], reverse=True)
    top_topics = "".join(
        f"<li><strong>{html.escape(topic)}</strong>: {count}</li>"
        for topic, count in result.topic_summary[:10]
    ) or "<li>No topics extracted</li>"

    article_blocks = "".join(
        (
            "<div style=\"margin:0 0 18px 0;padding:14px;border:1px solid #d9e2f2;border-radius:10px;background:#f8fbff;\">"
            f"<div style=\"font-size:12px;color:#4a5a70;margin-bottom:8px;\">{html.escape(item['source'])} | {html.escape(item['published_at'])}</div>"
            f"<div style=\"font-size:18px;font-weight:700;color:#0f172a;line-height:1.5;\">{html.escape(item.get('title_zh', '') or item['title'])}</div>"
            f"<div style=\"font-size:14px;color:#334155;line-height:1.7;margin-top:8px;\">{html.escape(item.get('summary_zh', '') or '暂无中文摘要')}</div>"
            f"<div style=\"margin-top:10px;font-size:13px;color:#64748b;\"><strong>EN Title:</strong> {html.escape(item['title'])}</div>"
            f"<div style=\"margin-top:6px;font-size:13px;color:#64748b;line-height:1.6;\"><strong>EN Summary:</strong> {html.escape(item['summary'])}</div>"
            f"<div style=\"margin-top:10px;\"><a href=\"{html.escape(item['link'])}\" style=\"color:#0b57d0;text-decoration:none;\">查看原文</a></div>"
            "</div>"
        )
        for item in sorted_items[:8]
    ) or "<div>No articles collected</div>"

    return f"""
<html>
  <body style="margin:0;padding:24px;background:#eef4fb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#0f172a;">
    <div style="max-width:900px;margin:0 auto;background:#ffffff;border-radius:16px;padding:28px;box-shadow:0 8px 30px rgba(15,23,42,0.08);">
    <h2 style="margin:0 0 10px 0;">国际热点双语简报</h2>
    <p style="margin:0 0 6px 0;color:#475569;">生成时间：{html.escape(result.generated_at.strftime("%Y-%m-%d %H:%M:%S %Z"))}</p>
    <p style="margin:0 0 6px 0;color:#475569;">文章数量：{len(result.items)}</p>
    <p style="margin:0 0 18px 0;color:#475569;">附件：{html.escape(result.output_path.name)}</p>
    <h3 style="margin:0 0 10px 0;">热点关键词</h3>
    <ul>{top_topics}</ul>
    <h3 style="margin:22px 0 12px 0;">重点新闻</h3>
    {article_blocks}
    </div>
  </body>
</html>
""".strip()


def send_digest_email(result: DigestResult, settings: EmailSettings) -> None:
    message = EmailMessage()
    message["Subject"] = f"International News Digest - {result.generated_at.strftime('%Y-%m-%d')}"
    message["From"] = settings.sender
    message["To"] = ", ".join(settings.recipients)
    message.set_content(
        "国际热点双语简报已生成并附在邮件中。\n"
        f"生成时间：{result.generated_at.strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
        f"文章数量：{len(result.items)}\n"
        "附件中包含中英文对照标题和摘要。\n"
        f"附件文件：{result.output_path.name}\n"
    )
    message.add_alternative(build_email_html(result), subtype="html")

    with result.output_path.open("rb") as file_obj:
        message.add_attachment(
            file_obj.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=result.output_path.name,
        )

    if settings.use_ssl:
        with smtplib.SMTP_SSL(settings.smtp_host, settings.smtp_port) as server:
            server.login(settings.smtp_user, settings.smtp_password)
            server.send_message(message)
    else:
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
            server.starttls()
            server.login(settings.smtp_user, settings.smtp_password)
            server.send_message(message)


def seconds_until_next_run(hour: int, minute: int, tz: ZoneInfo) -> float:
    now = datetime.now(tz)
    next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if next_run <= now:
        next_run += timedelta(days=1)
    return (next_run - now).total_seconds()


def run_scheduler(
    output_dir: Path,
    timezone_name: str,
    hour: int,
    minute: int,
    send_email: bool,
    default_recipient: str,
    translate_zh: bool,
) -> None:
    tz = ZoneInfo(timezone_name)
    print(f"[INFO] scheduler started, timezone={timezone_name}, run_at={hour:02d}:{minute:02d}")
    while True:
        wait_seconds = seconds_until_next_run(hour, minute, tz)
        next_time = datetime.now(tz) + timedelta(seconds=wait_seconds)
        print(f"[INFO] next run at {next_time.strftime('%Y-%m-%d %H:%M:%S %Z')}")
        time.sleep(wait_seconds)
        try:
            result = run_once(output_dir, timezone_name)
            if translate_zh:
                translation_settings = load_translation_settings()
                if translation_settings is None:
                    print("[WARN] translation skipped: missing translation provider credentials", file=sys.stderr)
                else:
                    translate_items_to_chinese(result.items, translation_settings)
                    write_excel(result.items, result.topic_summary, result.output_path, result.generated_at)
                    print(f"[INFO] Chinese translation completed using provider: {translation_provider_label(translation_settings)}")
            print(f"[INFO] Excel generated: {result.output_path}")
            if send_email:
                settings = load_email_settings(default_recipient)
                if settings is None:
                    print("[WARN] email skipped: missing SMTP configuration", file=sys.stderr)
                else:
                    send_digest_email(result, settings)
                    print(f"[INFO] Email sent to: {', '.join(settings.recipients)}")
        except Exception as exc:  # noqa: BLE001
            print(f"[ERROR] scheduled run failed: {exc}", file=sys.stderr)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Collect international media headlines and export to Excel.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated Excel files.")
    parser.add_argument("--timezone", default=DEFAULT_TIMEZONE, help="IANA timezone, e.g. Asia/Shanghai.")
    parser.add_argument("--hour", type=int, default=9, help="Hour for the daily scheduled run.")
    parser.add_argument("--minute", type=int, default=0, help="Minute for the daily scheduled run.")
    parser.add_argument("--daemon", action="store_true", help="Keep running and execute the task every day.")
    parser.add_argument("--send-email", action="store_true", help="Send the generated Excel by email after each run.")
    parser.add_argument("--recipient", default="363349082@qq.com", help="Default recipient email if NEWS_DIGEST_RECIPIENTS is not set.")
    parser.add_argument("--translate-zh", action="store_true", help="Translate titles and summaries into Simplified Chinese using the configured translation provider.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        if args.daemon:
            run_scheduler(
                args.output_dir,
                args.timezone,
                args.hour,
                args.minute,
                args.send_email,
                args.recipient,
                args.translate_zh,
            )
        else:
            result = run_once(args.output_dir, args.timezone)
            if args.translate_zh:
                translation_settings = load_translation_settings()
                if translation_settings is None:
                    raise RuntimeError("translation requested but translation provider credentials are not configured.")
                translate_items_to_chinese(result.items, translation_settings)
                write_excel(result.items, result.topic_summary, result.output_path, result.generated_at)
                print(f"[INFO] Chinese translation completed using provider: {translation_provider_label(translation_settings)}")
            print(f"[INFO] Excel generated: {result.output_path}")
            if args.send_email:
                settings = load_email_settings(args.recipient)
                if settings is None:
                    raise RuntimeError(
                        "email sending requested but SMTP configuration is incomplete. "
                        "Set NEWS_DIGEST_SMTP_HOST, NEWS_DIGEST_SMTP_PORT, NEWS_DIGEST_SMTP_USER, "
                        "NEWS_DIGEST_SMTP_PASSWORD and optionally NEWS_DIGEST_SENDER / NEWS_DIGEST_RECIPIENTS."
                    )
                send_digest_email(result, settings)
                print(f"[INFO] Email sent to: {', '.join(settings.recipients)}")
        return 0
    except KeyboardInterrupt:
        print("\n[INFO] scheduler stopped by user")
        return 130
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
